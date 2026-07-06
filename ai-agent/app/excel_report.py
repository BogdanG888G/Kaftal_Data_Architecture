"""Формирование полного Excel-отчёта."""
import io
import pandas as pd
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from database import db
from sql_builder import build_where
from config import config


TABLE = f"{config.CH_DATABASE}.{config.CH_TABLE}"


# Все доступные измерения
ALL_HIERARCHY_COLUMNS = {
    "retail_chain": "Сеть",
    "region_name": "Регион",
    "city_name": "Город",
    "store_format": "Формат магазина",
    "store_code": "Код магазина",
    "brand": "Бренд",
    "flavor": "Вкус",
    "weight_grams": "Граммовка",
    "manufacturer": "Производитель",
    "vendor": "Поставщик",
    "package_type": "Тип упаковки",
    "product_name": "Название товара",
    "chip_type": "Тип чипсов",
}


# Дефолтная иерархия
# ВАЖНО: address не в HIERARCHY, но идёт как метрика "Представленность в ТТ"
DEFAULT_HIERARCHY = [
    "retail_chain",
    "region_name",
    "city_name",
    "brand",
    "flavor",
    "weight_grams",
]


# Основные метрики
METRIC_COLUMNS = [
    ("tt_count", "Представленность в ТТ"),
    ("avg_cost", "Средняя себестоимость"),
    ("avg_sell", "Средняя цена продажи"),
    ("turnover_rub", "Товарооборот руб (с НДС)"),
    ("turnover_pcs", "Товарооборот шт"),
]


def build_report_sql(filters: dict, hierarchy_keys: list) -> str:
    """Строит SQL по иерархии + агрегатам."""
    where = build_where(filters)

    dim_cols_sql = []
    dim_cols_names = []

    for col_key in hierarchy_keys:
        if col_key not in ALL_HIERARCHY_COLUMNS:
            continue

        if col_key == "weight_grams":
            dim_cols_sql.append(
                f"toFloat64OrZero(toString({col_key})) AS {col_key}"
            )
        else:
            dim_cols_sql.append(f"ifNull({col_key}, 'Не указано') AS {col_key}")

        dim_cols_names.append(col_key)

    if not dim_cols_names:
        raise ValueError("Список иерархических колонок пуст")

    dim_select = ",\n    ".join(dim_cols_sql)
    dim_group = ", ".join(dim_cols_names)

    return f"""SELECT
    {dim_select},
    COUNT(DISTINCT address)                              AS tt_count,
    avg(average_cost_price)                              AS avg_cost,
    avg(average_sell_price)                              AS avg_sell,
    round(sum(sales_amount_rub), 2)                      AS turnover_rub,
    sum(sales_quantity)                                  AS turnover_pcs
FROM {TABLE}
{where}
GROUP BY {dim_group}
ORDER BY turnover_rub DESC"""


def clean_dataframe(df: pd.DataFrame, hierarchy_keys: list) -> pd.DataFrame:
    """Чистит DataFrame."""
    NULL_VALS = {"", "nan", "None", "NULL", "ᴺᵁᴸᴸ"}

    for col_key in hierarchy_keys:
        if col_key == "weight_grams":
            continue
        if col_key in df.columns:
            df[col_key] = df[col_key].astype(str).str.strip()
            df[col_key] = df[col_key].where(~df[col_key].isin(NULL_VALS), "Не указано")

    if "weight_grams" in df.columns:
        df["weight_grams"] = pd.to_numeric(df["weight_grams"], errors="coerce")

    # Фильтр адекватных цен
    if "turnover_rub" in df.columns and "turnover_pcs" in df.columns:
        real_price = df["turnover_rub"] / df["turnover_pcs"].replace(0, pd.NA)
        mask = (real_price >= 5) & (real_price <= 5000)
        df = df[mask].copy()

    return df


def remove_empty_hierarchy_columns(df: pd.DataFrame, hierarchy_keys: list) -> tuple:
    """Убирает колонки где всё пусто."""
    kept_columns = []

    for col_key in hierarchy_keys:
        if col_key not in df.columns:
            continue

        if col_key == "weight_grams":
            has_data = df[col_key].notna().any() and (df[col_key] > 0).any()
        else:
            unique_vals = df[col_key].dropna().unique()
            meaningful = [
                v for v in unique_vals
                if str(v).strip() not in ("", "Не указано")
            ]
            has_data = len(meaningful) > 0

        if has_data:
            kept_columns.append((col_key, ALL_HIERARCHY_COLUMNS[col_key]))
        else:
            df = df.drop(columns=[col_key])

    return df, kept_columns


def remove_empty_metric_columns(df: pd.DataFrame) -> pd.DataFrame:
    """Убирает метрики где всё NULL/0."""
    for metric_key, _ in METRIC_COLUMNS:
        if metric_key in df.columns:
            # tt_count / turnover_pcs = 0 значит нет данных
            if metric_key in ("tt_count", "turnover_pcs"):
                if df[metric_key].sum() == 0:
                    df = df.drop(columns=[metric_key])
            else:
                if df[metric_key].isna().all():
                    df = df.drop(columns=[metric_key])
    return df


def build_excel_bytes(filters: dict, excel_hierarchy: list = None) -> bytes:
    """
    Строит Excel-отчёт.
    
    Args:
        filters: dict с фильтрами.
        excel_hierarchy: список колонок из пожеланий пользователя.
                        Если пусто — используется DEFAULT_HIERARCHY.
    """
    # Определяем иерархию
    if excel_hierarchy:
        hierarchy_keys = [k for k in excel_hierarchy if k in ALL_HIERARCHY_COLUMNS]
        print(f"[EXCEL] User hierarchy: {hierarchy_keys}")
    else:
        hierarchy_keys = []

    # Если пусто — дефолт
    if not hierarchy_keys:
        hierarchy_keys = DEFAULT_HIERARCHY.copy()
        print(f"[EXCEL] Default hierarchy: {hierarchy_keys}")

    # 1. Запрос
    sql = build_report_sql(filters, hierarchy_keys)
    print(f"[EXCEL] SQL:\n{sql}")
    df = db.query(sql)
    print(f"[EXCEL] Rows: {len(df)}")

    if df.empty:
        raise ValueError("Нет данных по указанным фильтрам")

    # 2. Чистка
    df = clean_dataframe(df, hierarchy_keys)

    # 3. Убираем пустые иерархические колонки
    df, kept_hierarchy = remove_empty_hierarchy_columns(df, hierarchy_keys)

    # 4. Убираем пустые метрики
    df = remove_empty_metric_columns(df)

    # 5. Финальный DataFrame с русскими названиями
    rename_map = {}
    final_cols = []

    for col_key, col_name in kept_hierarchy:
        rename_map[col_key] = col_name
        final_cols.append(col_name)

    for col_key, col_name in METRIC_COLUMNS:
        if col_key in df.columns:
            rename_map[col_key] = col_name
            final_cols.append(col_name)

    df = df.rename(columns=rename_map)
    df = df[final_cols]

    # 6. Сортировка
    sort_by = [col_name for _, col_name in kept_hierarchy]
    if "Товарооборот руб (с НДС)" in df.columns:
        sort_by.append("Товарооборот руб (с НДС)")
        ascending = [True] * len(kept_hierarchy) + [False]
    else:
        ascending = [True] * len(kept_hierarchy)

    df = df.sort_values(sort_by, ascending=ascending).reset_index(drop=True)

    # 7. Excel
    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        sheet_name = "Отчёт"
        df.to_excel(writer, sheet_name=sheet_name, index=False)
        ws = writer.sheets[sheet_name]
        _apply_formatting(ws, df)

    buffer.seek(0)
    return buffer.getvalue()


def _apply_formatting(ws, df: pd.DataFrame):
    """Красивое форматирование."""
    header_fill = PatternFill("solid", fgColor="2F5496")
    header_font = Font(bold=True, size=11, color="FFFFFF")
    border_side = Side(style="thin", color="BFBFBF")
    thin_border = Border(
        left=border_side, right=border_side,
        top=border_side, bottom=border_side,
    )

    col_names = list(df.columns)

    # Шапка
    for col_idx, col_name in enumerate(col_names, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )
        cell.border = thin_border

    ws.row_dimensions[1].height = 32

    # Данные
    for row_idx in range(2, ws.max_row + 1):
        for col_idx, col_name in enumerate(col_names, start=1):
            cell = ws.cell(row=row_idx, column=col_idx)

            if col_name == "Товарооборот руб (с НДС)":
                cell.number_format = '#,##0.00 ₽'
                cell.alignment = Alignment(horizontal="right", vertical="center")
            elif col_name in ("Средняя себестоимость", "Средняя цена продажи"):
                cell.number_format = '#,##0.00'
                cell.alignment = Alignment(horizontal="right", vertical="center")
            elif col_name in ("Товарооборот шт", "Представленность в ТТ"):
                cell.number_format = '#,##0'
                cell.alignment = Alignment(horizontal="right", vertical="center")
            elif col_name == "Граммовка":
                cell.number_format = '0'
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center")

            cell.border = thin_border

    # Ширина колонок
    col_widths = {
        "Сеть": 14,
        "Регион": 24,
        "Город": 22,
        "Формат магазина": 18,
        "Код магазина": 16,
        "Бренд": 22,
        "Вкус": 30,
        "Граммовка": 12,
        "Производитель": 25,
        "Поставщик": 22,
        "Тип упаковки": 15,
        "Название товара": 45,
        "Тип чипсов": 20,
        "Представленность в ТТ": 20,
        "Средняя себестоимость": 22,
        "Средняя цена продажи": 22,
        "Товарооборот руб (с НДС)": 26,
        "Товарооборот шт": 18,
    }

    for col_idx, col_name in enumerate(col_names, start=1):
        letter = get_column_letter(col_idx)
        ws.column_dimensions[letter].width = col_widths.get(col_name, 18)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions